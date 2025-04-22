import json
import os
import pymsgbox
from openpyxl import load_workbook,Workbook
from pathlib import Path
from openpyxl.styles import Alignment
import datetime

#kataloguje trasy do pliku excela

def main():
    global data_to_store

    def store_data(file_path,row_cnt):

        with open(file_path, 'r') as file:
            data = json.load(file)

        file = '{}{}'.format(Path(__file__).parent,'''/trasy.xlsx''')
        wb=load_workbook(file)
        ws = wb['trasy']

        row_cnt+=2

        ws.cell(row=row_cnt,column=1).value=str(row_cnt-1)

        # for i,j in enumerate(data_to_store):
        #     if j=='EstimatedTime':
        #         ws.cell(row=row_cnt,column=i+2).value=str(datetime.timedelta(seconds=data["RouteInfo"][j]))
        #     else:
        #         ws.cell(row=row_cnt,column=i+2).value=data["RouteInfo"][j]

        ws.cell(row=row_cnt,column=2).value=data["RouteInfo"]['Name']
        ws.cell(row=row_cnt,column=3).value=data["RouteInfo"]['Rating']
        ws.cell(row=row_cnt,column=4).value=data["RouteInfo"]['Length']/1000
        ws.cell(row=row_cnt,column=5).value=str(datetime.timedelta(seconds=data["RouteInfo"]['EstimatedTime']))
        ws.cell(row=row_cnt,column=6).value=data["RouteInfo"]['Upclimb']
        ws.cell(row=row_cnt,column=7).value=data["RouteInfo"]['MaxSlope']
        ws.cell(row=row_cnt,column=8).value=data["RouteInfo"]['AvgSlope']
        ws.cell(row=row_cnt,column=9).value=data["RouteInfo"]['Id']

        ws.column_dimensions['A'].width=3
        ws.column_dimensions['B'].width=70


        wb.save(file)
        wb.close()
        
    input_dir = os.getcwd()

    cnt=0
    error_files=''

    def create_file():
        # stworzenie pliku z trasami
        filepath = '{}{}'.format(Path(__file__).parent,'''/trasy.xlsx''')
        wb = Workbook()
        wb.create_sheet('trasy')
        ws = wb['trasy']

        for i,j in enumerate(data_to_store):
            ws.cell(row=1,column=i+2).value=j

        wb.save(filepath)       
        wb.save(filepath)


    data_to_store=['Name','Rating','Length','EstimatedTime','Upclimb','MaxSlope','AvgSlope','Id']

    create_file()

    for i,_,k in os.walk(input_dir):
        for l in k:
            if l.endswith('.json'):
                    try:
                        file_path='{}\\{}'.format(i,l)
                        store_data(file_path,cnt)
                        cnt+=1
                    except:
                        error_files+=l.split('''\\''')[-1]+'\n'
                    


    pymsgbox.alert(text=f'skatalogowane trasy :{cnt}szt.\nBłędne pliki:\n'+error_files)

if __name__ == "__main__":
    main()
    pass


# # Open and read the JSON file
# with open('35300.json', 'r') as file:
#     data = json.load(file)

# # # # Print the data
# # print(data["RouteInfo"])

# for i,j in data["RouteInfo"].items():
#     print(i+'     :     '+str(j))
